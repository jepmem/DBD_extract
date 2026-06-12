"""
scrape_franchise_v3.py  —  แก้ให้ตรงกับโครงสร้าง HTML จริง

ข้อมูลจาก HTML จริง:
  LIST PAGE:
    - card selector:  div.category-list a   (href เป็น relative เช่น "freshy?csrt=...")
    - ชื่อใน card:    figcaption h2
    - ปุ่มโหลดเพิ่ม: button.loadmore-btn   (มีทุก category-tab)

  DETAIL PAGE:
    - ชื่อ franchise: h1 แรกในหน้า
    - label จริง:    ชื่อกิจการ, หมวดหมู่, ประเภท, งบลงทุนต่ำสุด/สูงสุด,
                     ค่าแรกเข้า, ค่ารายปี, จำนวนสาขา, ชื่อผู้ติดต่อ,
                     ที่อยู่, โทรศัพท์, โทรศัพท์มือถือ, แฟกซ์,
                     อีเมล์  (มีวรรณยุกต์!), เว็บไซต์
"""

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import pandas as pd
import re
import time
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL    = "https://franchise.dbd.go.th"
LIST_URL    = f"{BASE_URL}/th/franchise-category/list"
OUTPUT_FILE = "franchise_data_v3.xlsx"
MAX_ITEMS   = 200   # ตั้งเป็น 0 เพื่อดึงทั้งหมด

# ──────────────── helpers ────────────────

def ef(body_text, label):
    """extract_field: ดึงค่าหลัง label"""
    m = re.search(rf'{re.escape(label)}\s*[:\uf03a]?\s*(.+?)(?=\n|$)', body_text)
    return m.group(1).strip() if m else ""

def scrape_detail(page, url, retries=3):
    for attempt in range(retries):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(5000)
            body = page.inner_text("body")
            if len(body) < 100:
                raise Exception("body too short")

            # ชื่อ franchise จาก h1 แรก (ก่อน navigation)
            name = ""
            try:
                # หา h1 ที่ไม่ใช่ header/nav
                for el in page.locator("h1").all():
                    txt = el.inner_text().strip()
                    # กรองออก h1 ที่เป็น logo/header
                    if txt and len(txt) < 150 and "กรมพัฒนา" not in txt and "Department" not in txt:
                        name = txt
                        break
            except:
                pass

            invest_min = ef(body, "งบลงทุนต่ำสุด")
            invest_max = ef(body, "งบลงทุนสูงสุด")
            if invest_min:
                invest_range = f"{invest_min} - {invest_max}" if invest_max else invest_min
            else:
                invest_range = ""

            line_m = re.search(r'Line\s*ID\s*[:\s]\s*(\S+)', body, re.IGNORECASE)

            return {
                "ชื่อ Franchise":  name,
                "ชื่อกิจการ":      ef(body, "ชื่อกิจการ"),
                "หมวดหมู่":        ef(body, "หมวดหมู่"),
                "ประเภท":          ef(body, "ประเภท"),
                "งบลงทุน (บาท)":   invest_range,
                "ค่าแรกเข้า":      ef(body, "ค่าแรกเข้า"),
                "ค่ารายปี":        ef(body, "ค่ารายปี"),
                "จำนวนสาขา":       ef(body, "จำนวนสาขา"),
                "ผู้ติดต่อ":        ef(body, "ชื่อผู้ติดต่อ"),
                "ที่อยู่":          ef(body, "ที่อยู่"),
                "โทรศัพท์":         ef(body, "โทรศัพท์"),
                "มือถือ":           ef(body, "โทรศัพท์มือถือ"),
                "แฟกซ์":            ef(body, "แฟกซ์"),
                "อีเมล์":           ef(body, "อีเมล์"),   # ← วรรณยุกต์!
                "เว็บไซต์":         ef(body, "เว็บไซต์"),
                "Line ID":          line_m.group(1) if line_m else "",
                "URL":              url,
                "สถานะ":            "OK",
            }

        except Exception as e:
            print(f"  attempt {attempt+1}/{retries} failed: {e}")
            if attempt < retries - 1:
                time.sleep(5)

    return {k: "" for k in [
        "ชื่อ Franchise","ชื่อกิจการ","หมวดหมู่","ประเภท","งบลงทุน (บาท)",
        "ค่าแรกเข้า","ค่ารายปี","จำนวนสาขา","ผู้ติดต่อ","ที่อยู่",
        "โทรศัพท์","มือถือ","แฟกซ์","อีเมล์","เว็บไซต์","Line ID"
    ]} | {"URL": url, "สถานะ": "ERROR: max retries"}


def collect_urls_with_load_more(page, max_items=200):
    """
    เปิดหน้า list แล้วกดปุ่ม loadmore-btn ทีละ category
    URL ใน card เป็น relative เช่น 'freshy?csrt=...'
    ต้อง normalise เป็น https://franchise.dbd.go.th/franchise/freshy
    """
    print(f"เปิด: {LIST_URL}")
    page.goto(LIST_URL, wait_until="domcontentloaded", timeout=45000)
    page.wait_for_timeout(5000)

    urls = {}   # url → ชื่อ franchise (จาก card)

    def harvest():
        """ดูด URL + ชื่อจาก card ที่โหลดแล้ว"""
        for card in page.locator("div.category-list a").all():
            try:
                href = card.get_attribute("href") or ""
                if not href or href.startswith("#") or href.startswith("javascript"):
                    continue
                # normalise
                clean = href.split("?")[0].strip("/")
                if clean.startswith("http"):
                    full_url = clean
                elif clean.startswith("/"):
                    full_url = f"{BASE_URL}{clean}"
                else:
                    full_url = f"{BASE_URL}/franchise/{clean}"

                if full_url not in urls:
                    # ดึงชื่อจาก h2 ใน figcaption
                    try:
                        name = card.locator("figcaption h2").first.inner_text().strip()
                    except:
                        name = ""
                    urls[full_url] = name
            except:
                pass

    harvest()
    print(f"  initial: {len(urls)} URLs")

    round_num = 0
    while True:
        if max_items and len(urls) >= max_items:
            break

        # กด loadmore-btn ทุกอันที่ยัง visible
        btns = page.locator("button.loadmore-btn").all()
        if not btns:
            print("  ไม่พบปุ่มโหลดเพิ่มเติม")
            break

        clicked = 0
        for btn in btns:
            try:
                if btn.is_visible():
                    btn.scroll_into_view_if_needed()
                    btn.click()
                    clicked += 1
            except:
                pass

        if clicked == 0:
            print("  ไม่มีปุ่มที่กดได้ — หยุด")
            break

        page.wait_for_timeout(3000)
        before = len(urls)
        harvest()
        round_num += 1
        print(f"  กดครั้งที่ {round_num} ({clicked} ปุ่ม): {len(urls)} URLs (เพิ่ม {len(urls)-before})")

        if len(urls) == before:
            print("  ไม่มี URL ใหม่ — หยุด")
            break

    result = list(urls.items())   # [(url, name), ...]
    if max_items:
        result = result[:max_items]
    return result


# ──────────────── save ────────────────

def save_excel(data, filename=OUTPUT_FILE):
    df = pd.DataFrame(data)
    col_widths = {
        "A":28,"B":30,"C":15,"D":20,"E":25,"F":18,
        "G":15,"H":15,"I":25,"J":45,"K":18,"L":18,
        "M":15,"N":30,"O":35,"P":20,"Q":50,"R":12
    }
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Franchise")
        ws = writer.sheets["Franchise"]
        hfill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, (col, w) in enumerate(col_widths.items(), 1):
            ws.column_dimensions[col].width = w
        rfill = PatternFill("solid", fgColor="FFE0E0")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
            if str(row[-1].value or "").startswith("ERROR"):
                for cell in row:
                    cell.fill = rfill
        ws.freeze_panes = "A2"
    ok  = sum(1 for d in data if d.get("สถานะ") == "OK")
    err = len(data) - ok
    print(f"✅ บันทึก {filename}: {ok} OK, {err} ERROR จาก {len(data)} รายการ")


# ──────────────── main ────────────────

def scrape():
    data = []
    with sync_playwright() as p:
        ctx = p.chromium.launch(headless=False).new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )
        page = ctx.new_page()

        # ── ขั้น 1: รวบรวม URL ──
        pairs = collect_urls_with_load_more(page, max_items=MAX_ITEMS)
        print(f"\nจะ scrape {len(pairs)} รายการ\n")

        # ── ขั้น 2: scrape ทีละหน้า ──
        for i, (url, card_name) in enumerate(pairs, 1):
            print(f"[{i}/{len(pairs)}] {url}")
            row = scrape_detail(page, url)

            # ถ้า h1 ไม่เจอ fallback ใช้ชื่อจาก card
            if not row["ชื่อ Franchise"] and card_name:
                row["ชื่อ Franchise"] = card_name

            data.append(row)
            status = "✓" if row["สถานะ"] == "OK" else "✗"
            print(f"  {status} {row['ชื่อ Franchise'] or '(ไม่มีชื่อ)'}")

            if i % 20 == 0:
                save_excel(data, f"franchise_checkpoint_{i}.xlsx")
            time.sleep(1.5)

        # ── ขั้น 3: retry ERROR ──
        errors = [i for i, d in enumerate(data) if str(d.get("สถานะ","")).startswith("ERROR")]
        if errors:
            print(f"\n🔄 retry {len(errors)} รายการ...")
            for idx in errors:
                time.sleep(3)
                data[idx] = scrape_detail(page, data[idx]["URL"], retries=2)

        ctx.browser.close()

    save_excel(data)

if __name__ == "__main__":
    scrape()