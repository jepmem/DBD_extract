"""
DBD DataWarehouse Full Scraper v7
====================================
แก้ไขจาก v6 — ใช้ Apify แทน Google CSE API:

  Phase 1 (DBD scrape):
    - ใช้ real Chrome profile → Google เห็นว่าเป็น user จริง มีประวัติ cookies

  Phase 2 (phone lookup):
    - ใช้ Apify Google Search scraper แทน browser
    - ไม่มี CAPTCHA เด็ดขาด
    - ฟรี $5/month credit (~1,000-2,000 searches)

  การตั้งค่า (ทำครั้งเดียว):
    1. ไปที่ https://console.apify.com/sign-up สมัครฟรี
    2. ไปที่ Settings → Integrations → API tokens → copy Personal API token
    3. ใส่ค่าใน APIFY_API_TOKEN ด้านล่าง

  Chrome profile path (แก้ให้ตรงกับเครื่องตัวเอง):
    macOS  : /Users/YOUR_NAME/Library/Application Support/Google/Chrome/Default
    Windows: C:\\Users\\YOUR_NAME\\AppData\\Local\\Google\\Chrome\\User Data\\Default
    Linux  : /home/YOUR_NAME/.config/google-chrome/Default
"""

from __future__ import annotations

import asyncio
import re
import json
import random
import sys
import platform
import os
from datetime import datetime
from pathlib import Path

import httpx
import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, Page


# ══════════════════════════════════════════
# CONFIG  ← แก้บรรทัดนี้
# ══════════════════════════════════════════
APIFY_API_TOKEN   = os.environ.get("APIFY_API_TOKEN", "")    # Personal API token: set via env var
CHROME_PROFILE    = ""                          # ถ้าเว้นว่าง จะใช้ default ตาม OS อัตโนมัติ

DBD_HOME          = "https://datawarehouse.dbd.go.th"
DBD_DELAY         = 0.35
GOOGLE_MIN        = 2.0    # Apify ช้ากว่า API ตรง แต่เร็วกว่า browser
GOOGLE_MAX        = 4.0
SAVE_EVERY        = 50
MAX_GOOGLE_PER_SESSION = 150   # Apify free tier: ~$5/month ≈ 1,000-2,000 searches
PROFILE_DELAY     = 0.12


# ── Auto-detect Chrome profile path ────────────────────────────────────────
def _default_chrome_profile() -> str:
    if CHROME_PROFILE:
        return CHROME_PROFILE
    system = platform.system()
    home = Path.home()
    if system == "Darwin":
        return str(home / "Library/Application Support/Google/Chrome/Default")
    if system == "Windows":
        return str(home / "AppData/Local/Google/Chrome/User Data/Default")
    return str(home / ".config/google-chrome/Default")

def _chrome_exe() -> str | None:
    """Return path to real Chrome executable, or None to let Playwright use Chromium."""
    system = platform.system()
    candidates = {
        "Darwin":  ["/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"],
        "Windows": [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        ],
        "Linux":   ["/usr/bin/google-chrome", "/usr/bin/chromium-browser"],
    }.get(system, [])
    for p in candidates:
        if Path(p).exists():
            return p
    return None


# ══════════════════════════════════════════
# BLOCKED DOMAINS
# ══════════════════════════════════════════
BLOCKED_DOMAINS = [
    "dataforthai.com",
    "dbd.go.th",
    "datawarehouse.dbd.go.th",
    "thaidbsearch.com",
]


async def dismiss_blocking_modals(page: Page) -> None:
    """Close DBD popups that can intercept clicks on the search controls."""
    selectors = [
        "#warningModal button:has-text('ตกลง')",
        "#warningModal button:has-text('ยอมรับ')",
        "#warningModal button:has-text('ปิด')",
        "#warningModal .btn-close",
        ".modal.show button:has-text('ตกลง')",
        ".modal.show button:has-text('ยอมรับ')",
        ".modal.show button:has-text('ปิด')",
        ".modal.show .btn-close",
    ]
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            if await locator.count() and await locator.is_visible(timeout=1000):
                await locator.click(timeout=3000)
                await page.wait_for_timeout(700)
                return
        except Exception:
            continue

    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(500)
    except Exception:
        pass


# ══════════════════════════════════════════
# PHONE EXTRACTOR  (v5 — unchanged)
# ══════════════════════════════════════════
PHONE_RE = re.compile(
    r"(?<!\d)"
    r"(0[2-9]\d)"
    r"([-.\s]?)"
    r"(\d{3,4})"
    r"\2"
    r"(\d{3,4})"
    r"(?!\d)"
)

def _is_real_phone(digits: str) -> bool:
    if len(digits) not in (9, 10):
        return False
    if not digits.startswith("0"):
        return False
    prefix2 = digits[:2]
    prefix3 = digits[:3]
    if len(digits) == 10 and prefix2 in ("06", "08", "09"):
        if digits[2:] in ("00000000", "11111111", "12345678"):
            return False
        return True
    if prefix2 == "02" and len(digits) == 9:
        return True
    if prefix2 in ("03", "04", "05", "07") and len(digits) == 10:
        return True
    if prefix3 in ("077", "078", "079") and len(digits) == 10:
        return True
    return False

def _format_phone(digits: str) -> str:
    if len(digits) == 10:
        return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 9:
        if digits[:2] == "02":
            return f"{digits[:2]}-{digits[2:5]}-{digits[5:]}"
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return digits

def extract_phones(text: str) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for m in PHONE_RE.finditer(text):
        digits = re.sub(r"[^\d]", "", m.group(0))
        if digits in seen or not _is_real_phone(digits):
            continue
        seen.add(digits)
        result.append(_format_phone(digits))
    return result


def _profile_url(registration_no: str) -> str:
    digits = re.sub(r"\D", "", registration_no or "")
    if len(digits) >= 13:
        profile_id = f"5{digits}"
    else:
        profile_id = digits
    return f"{DBD_HOME}/company/profile/{profile_id}"


def _extract_head_office_address(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for i, line in enumerate(lines):
        if line == "ที่ตั้งสำนักงานแห่งใหญ่" and i + 1 < len(lines):
            return lines[i + 1]
    return ""


def _extract_objectives(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    stop_words = {
        "ประเภทธุรกิจ",
        "ประเภทธุรกิจที่ส่งงบการเงินปีล่าสุด",
        "หมายเหตุ",
        "รายชื่อกรรมการ",
        "กรรมการลงชื่อผูกพัน",
        "Website",
        "ให้น้อง SUPER ช่วยคุณนะครับ",
    }
    objectives = []
    for i, line in enumerate(lines):
        if line != "วัตถุประสงค์":
            continue
        chunk = []
        for next_line in lines[i + 1:]:
            if next_line in stop_words or next_line.startswith("ข้อมูลนิติบุคคล"):
                break
            chunk.append(next_line)
        if chunk:
            objectives.append(" ".join(chunk))
    return " ".join(objectives)


def _extract_district(address: str) -> str:
    patterns = [
        r"(?:^|\s)เขต\s*([^\s,]+)",
        r"(?:^|\s)อำเภอ\s*([^\s,]+)",
        r"(?:^|\s)อ\.\s*([^\s,]+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, address)
        if m:
            return m.group(1).strip()
    return ""


def classify_food_business(name: str, business_name: str, objective: str) -> str:
    # Do not let broad DBD TSIC labels like "ภัตตาคาร/ร้านอาหาร" force every row
    # into "ภัตตาคาร"; the subtype should come from the profile objective and name.
    context = " ".join([name or "", objective or ""]).lower()

    restaurant_keywords = [
        "ภัตตาคาร",
        "restaurant",
        "restaurants",
        "เรสเตอรอง",
        "เรสเทอรอง",
    ]
    cafe_keywords = [
        "คาเฟ่",
        "คาเฟ",
        "ร้านกาแฟ",
        "กาแฟ",
        "coffee",
        "cafe",
        "café",
    ]

    if any(keyword in context for keyword in restaurant_keywords):
        return "ภัตตาคาร"
    if any(keyword in context for keyword in cafe_keywords):
        return "คาเฟ่"
    return "ร้านอาหาร"


async def _profile_body_text(page: Page) -> str:
    text = ""
    for _ in range(10):
        text = await page.locator("body").inner_text(timeout=10000)
        if "ที่ตั้งสำนักงานแห่งใหญ่" in text:
            return text
        await page.wait_for_timeout(1000)
    return text


async def _wait_for_juristic_rows(page: Page, timeout_ms: int = 30000) -> bool:
    try:
        await page.wait_for_function(
            """() => {
                const row = document.querySelector('tbody tr');
                return row && !row.innerText.includes('Loading...');
            }""",
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


async def _wait_no_loader(page: Page, timeout_ms: int = 30000) -> None:
    try:
        await page.locator(".loader-overlay-full").wait_for(state="hidden", timeout=timeout_ms)
    except Exception:
        pass


DBD_STATUS_FILTERS = ["ยังดำเนินกิจการอยู่", "ฟื้นฟู", "ควบ", "แปรสภาพ"]


async def apply_status_filters(page: Page) -> None:
    print("  [filter] เลือกสถานะ active บนเว็บ DBD...")
    await _wait_no_loader(page)

    try:
        await page.get_by_role("button", name="ตัวกรองข้อมูลเพิ่มเติม").click(timeout=10000)
        await page.wait_for_timeout(800)
    except Exception:
        pass

    await _wait_no_loader(page)
    status_box = page.locator(".filter-advanced .box").filter(
        has=page.locator("h5", has_text="สถานะ")
    ).first
    await status_box.locator(".multiselect").click(timeout=15000)
    await page.wait_for_timeout(500)

    for status in DBD_STATUS_FILTERS:
        await page.get_by_text(status, exact=True).last.click(timeout=15000)
        await page.wait_for_timeout(250)

    await _wait_no_loader(page)
    async with page.expect_response(
        lambda r: "/api/v1/company-profiles/infos" in r.url and r.request.method == "POST",
        timeout=30000,
    ):
        await page.locator(".filter-advanced").get_by_role(
            "button", name=re.compile(r"ค้นหาข้อมูล")
        ).click(timeout=15000)

    await _wait_for_juristic_rows(page, timeout_ms=30000)
    await page.wait_for_timeout(1200)
    print(f"      ✓ กรองสถานะ: {', '.join(DBD_STATUS_FILTERS)}")


async def select_capital_sort(page: Page) -> None:
    try:
        sort_dropdown = page.locator("select").last
        await sort_dropdown.select_option(value="capAmt", timeout=5000)
        await page.wait_for_load_state("domcontentloaded")
        await _wait_for_juristic_rows(page, timeout_ms=10000)
        await page.wait_for_timeout(1000)
        print("      ✓ เรียงตามทุนจดทะเบียนแล้ว")
    except Exception as e:
        print(f"      ⚠️  ไม่พบ dropdown การเรียง (ข้าม): {e}")


async def go_to_result_page(page: Page, page_no: int) -> None:
    if page_no <= 1:
        return
    page_input = page.locator("input.form-control.numeric[type='number']").first
    await page_input.fill(str(page_no))
    try:
        async with page.expect_response(
            lambda r: "/api/v1/company-profiles/infos" in r.url and r.request.method == "POST",
            timeout=30000,
        ):
            await page_input.press("Enter")
    except Exception:
        await page_input.press("Enter")
    await _wait_for_juristic_rows(page, timeout_ms=30000)
    await page.wait_for_timeout(1000)


async def refresh_search_page(page: Page, keyword: str, page_no: int) -> None:
    print(f"↻ refresh หน้า {page_no} ", end="", flush=True)
    direct_url = f"{DBD_HOME}/juristic/searchInfo?keyword={keyword}&v=biz"
    await page.goto(direct_url, wait_until="domcontentloaded", timeout=60000)
    await dismiss_blocking_modals(page)
    await _wait_for_juristic_rows(page, timeout_ms=30000)
    await select_capital_sort(page)
    await apply_status_filters(page)
    await go_to_result_page(page, page_no)


# ══════════════════════════════════════════
# NAVIGATE TO JURISTIC LIST  (unchanged)
# ══════════════════════════════════════════
async def navigate_to_list(page: Page, keyword: str) -> bool:
    print(f"\n{'═'*60}")
    print(f"  NAVIGATE — keyword={keyword}")
    print(f"{'═'*60}")

    print("  [1/7] เปิดหน้าแรก...")
    direct_url = f"{DBD_HOME}/juristic/searchInfo?keyword={keyword}&v=biz"
    try:
        await page.goto(direct_url, wait_until="domcontentloaded", timeout=60000)
        await dismiss_blocking_modals(page)
        if await _wait_for_juristic_rows(page, timeout_ms=30000):
            print("      ✓ เข้า list ด้วย URL ตรง")
            print("  [7/7] เลือก 'ทุนจดทะเบียน (มาก-น้อย)'...")
            await select_capital_sort(page)
            await apply_status_filters(page)
            print("  ✅ นำทางสำเร็จ\n")
            return True
    except Exception as e:
        print(f"      ⚠️  URL ตรงไม่สำเร็จ ใช้วิธีคลิกฟอร์มแทน: {e}")

    await page.goto(DBD_HOME, wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2000)
    await dismiss_blocking_modals(page)

    print("  [2/7] คลิก 'ค้นหาแบบมีเงื่อนไข'...")
    for attempt in range(3):
        await page.get_by_text("ค้นหาแบบมีเงื่อนไข", exact=False).first.click()
        await page.wait_for_timeout(1500)
        if await page.locator("select").count():
            break
        if attempt < 2:
            await page.reload(wait_until="domcontentloaded", timeout=60000)
            await page.wait_for_timeout(2000)
            await dismiss_blocking_modals(page)

    print("  [3/7] เลือก 'รหัสประเภทธุรกิจ'...")
    await page.locator("select").first.wait_for(state="attached", timeout=15000)
    sel = page.locator("select").first
    try:
        await sel.select_option(label="รหัสประเภทธุรกิจ")
    except Exception:
        await sel.click()
        await page.wait_for_timeout(400)
        await page.get_by_text("รหัสประเภทธุรกิจ", exact=True).first.click()
    await page.wait_for_timeout(500)

    print(f"  [4/7] กรอก '{keyword}' → ตกลง...")
    await page.locator("input[type='text'], input[type='search']").last.fill(keyword)
    await page.wait_for_timeout(300)
    await page.get_by_role("button", name=re.compile(r"ตกลง|ค้นหา", re.I)).first.click()
    await page.wait_for_load_state("domcontentloaded")
    await page.wait_for_timeout(2000)

    print("  [5/7] คลิกแถวผลลัพธ์...")
    await page.locator(f"tr:has-text('{keyword}')").first.click()
    await page.wait_for_load_state("domcontentloaded")
    await page.wait_for_timeout(2000)

    print("  [6/7] กด 'แสดงรายชื่อนิติบุคคล'...")
    await page.get_by_text(re.compile(r"แสดงรายชื่อนิติบุคคล", re.I)).first.click()
    await page.wait_for_load_state("domcontentloaded")
    await page.wait_for_timeout(2500)

    print("  [7/7] เลือก 'ทุนจดทะเบียน (มาก-น้อย)'...")
    await select_capital_sort(page)
    await apply_status_filters(page)
    print("  ✅ นำทางสำเร็จ\n")
    return True


# ══════════════════════════════════════════
# PARSE TABLE  (unchanged)
# ══════════════════════════════════════════
ACTIVE_STATUS_KEYWORDS = [
    "ยังดำเนินกิจการอยู่",
    "ดำเนินการอยู่",
    "ฟื้นฟู",
    "ควบ",
    "แปรสภาพ",
]


def is_active_company_status(status: str) -> bool:
    return any(keyword in (status or "") for keyword in ACTIVE_STATUS_KEYWORDS)


def filter_active_companies(companies: list[dict]) -> list[dict]:
    return [company for company in companies if is_active_company_status(company.get("สถานะ", ""))]


async def parse_page_html(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "lxml")
    companies = []
    table = soup.select_one("table")
    if not table:
        return companies

    headers_raw = [th.get_text(strip=True) for th in table.select("thead th, tr:first-child th")]
    rows = table.select("tbody tr") or table.select("tr")[1:]

    for row in rows:
        cells = [td.get_text(" ", strip=True) for td in row.find_all("td")]
        if len(cells) < 2:
            continue

        if headers_raw and len(headers_raw) == len(cells):
            record = dict(zip(headers_raw, cells))
        else:
            record = {}
            for c in cells:
                c = c.strip()
                if re.match(r"^0\d{12}$", re.sub(r"[\s\-]", "", c)):
                    record["เลขทะเบียน"] = c
                elif any(k in c for k in ["จำกัด", "หจก", "บริษัท", "ห้างหุ้นส่วน"]) and len(c) > 4:
                    record.setdefault("ชื่อบริษัท", c)
                elif c in ["บริษัทจำกัด", "บริษัทมหาชนจำกัด", "ห้างหุ้นส่วนจำกัด", "ห้างหุ้นส่วนสามัญนิติบุคคล"]:
                    record["ประเภทนิติบุคคล"] = c
                elif c in ["ยังดำเนินกิจการอยู่", "เลิก", "ร้าง", "ถูกขีดชื่อ", "ยุบเลิก"]:
                    record["สถานะ"] = c
                elif re.match(r"^\d{4,6}$", c):
                    record.setdefault("รหัสประเภทธุรกิจ", c)
                elif len(c) > 8 and "ชื่อประเภทธุรกิจ" not in record and "ชื่อบริษัท" in record:
                    record["ชื่อประเภทธุรกิจ"] = c

        name = record.get("ชื่อบริษัท") or record.get("ชื่อนิติบุคคล") or ""
        if not name:
            for k, v in record.items():
                if "ชื่อ" in k and v:
                    name = v
                    break

        if name:
            companies.append({
                "เลขทะเบียนนิติบุคคล":           record.get("เลขทะเบียนนิติบุคคล", record.get("เลขทะเบียน", "")),
                "ชื่อนิติบุคคล":                  name,
                "ประเภทนิติบุคคล":                record.get("ประเภทนิติบุคคล", record.get("ประเภท", "")),
                "สถานะ":                          record.get("สถานะ", ""),
                "รหัสประเภทธุรกิจ":               record.get("รหัสประเภทธุรกิจ", record.get("รหัสประเภท", "")),
                "ชื่อประเภทธุรกิจ":               record.get("ชื่อประเภทธุรกิจ", record.get("ประเภทธุรกิจ", "")),
                "ประเภทธุรกิจย่อย":                "",
                "จังหวัด":                         record.get("จังหวัด", ""),
                "เขต/อำเภอ":                       record.get("เขต/อำเภอ", ""),
                "ทุนจดทะเบียน (บาท)":             record.get("ทุนจดทะเบียน (บาท)", ""),
                "รายได้รวม (บาท)":                record.get("รายได้รวม (บาท)", ""),
                "กำไร (ขาดทุน) สุทธิ (บาท)":      record.get("กำไร (ขาดทุน) สุทธิ (บาท)", ""),
                "สินทรัพย์รวม (บาท)":             record.get("สินทรัพย์รวม (บาท)", ""),
                "ส่วนของผู้ถือหุ้น (บาท)":        record.get("ส่วนของผู้ถือหุ้น (บาท)", ""),
                "เบอร์โทร":                       "",
                "เว็บไซต์":                        "",
            })

    return companies


# ══════════════════════════════════════════
# SCRAPE ALL PAGES  (unchanged)
# ══════════════════════════════════════════
async def _get_total_pages(page: Page) -> tuple[int, int]:
    html = await page.content()
    text = BeautifulSoup(html, "lxml").get_text(" ")
    total_items = 0
    total_pages = 0
    m = re.search(r"([\d,]+)\s*ราย(?:การ|นิติบุคคล)?", text)
    if m:
        total_items = int(m.group(1).replace(",", ""))
    m2 = re.search(r"หน้า\s*\d+\s*/\s*([\d,]+)", text)
    if m2:
        total_pages = int(m2.group(1).replace(",", ""))
    if total_pages == 0 and total_items > 0:
        total_pages = (total_items + 9) // 10
    return total_items, total_pages


async def _click_next(page: Page) -> bool:
    first_row_before = await page.locator("tbody tr").first.inner_text()
    btn = page.locator("button.pagingBtn").last
    if await btn.count() == 0:
        return False
    dis = await btn.get_attribute("disabled")
    cls = await btn.get_attribute("class") or ""
    if dis is not None or "disabled" in cls:
        return False
    try:
        async with page.expect_response(
            lambda r: "/api/v1/company-profiles/infos" in r.url and r.request.method == "POST",
            timeout=20000,
        ):
            await btn.click()
    except Exception:
        # The UI occasionally updates from cache or resolves slower than the response hook.
        # Fall through to the DOM-level row-change wait below.
        pass
    try:
        await page.wait_for_function(
            f"""() => {{
                const row = document.querySelector('tbody tr');
                return row
                    && !row.innerText.includes('Loading...')
                    && row.innerText.trim() !== {json.dumps(first_row_before.strip())};
            }}""",
            timeout=20000,
        )
    except Exception:
        await page.wait_for_timeout(3000)
    await _wait_for_juristic_rows(page, timeout_ms=15000)
    return True


async def refresh_current_search_page(page: Page) -> None:
    print("↻ refresh ", end="", flush=True)
    await page.reload(wait_until="domcontentloaded", timeout=60000)
    await _wait_for_juristic_rows(page, timeout_ms=30000)
    await page.wait_for_timeout(1000)


async def scrape_dbd_all(
    page: Page,
    keyword: str,
    max_pages: int = None,
    start_page: int = 1,
    initial_companies: list[dict] = None,
) -> list[dict]:
    total_items, total_pages = await _get_total_pages(page)
    if max_pages is None:
        max_pages = total_pages if total_pages > 0 else 9999
    if start_page > 1:
        await go_to_result_page(page, start_page)

    print(f"\n{'═'*60}")
    print(f"  PHASE 1 — Scrape รายชื่อนิติบุคคล")
    print(f"  พบ ~{total_items:,} ราย | {total_pages:,} หน้า")
    print(f"  จะดึงหน้า {start_page:,} ถึง {max_pages:,}")
    print(f"{'═'*60}\n")

    all_cos: list[dict] = list(initial_companies or [])
    for p in range(start_page, max_pages + 1):
        pct = p / max_pages * 100
        print(f"  [{p:>5}/{max_pages}] {pct:4.0f}%  ", end="", flush=True)
        rows = []
        for retry in range(12):
            await _wait_for_juristic_rows(page, timeout_ms=15000)
            html = await page.content()
            rows = await parse_page_html(html)
            if rows:
                break
            if retry in (3, 7):
                await refresh_search_page(page, keyword, p)
            await page.wait_for_timeout(2500)
        if rows:
            active_rows = filter_active_companies(rows)
            skipped = len(rows) - len(active_rows)
            all_cos.extend(active_rows)
            print(f"✓ {len(active_rows)}/{len(rows)} ราย active  ข้าม {skipped}  (รวม {len(all_cos):,})")
        else:
            raise RuntimeError(f"ไม่พบข้อมูลในหน้า {p} หลัง retry แล้ว — หยุดเพื่อกันข้อมูลขาด")
        if p % SAVE_EVERY == 0:
            _save_json(all_cos, f"checkpoint_{keyword}_p{p}.json")
            print(f"         💾 checkpoint ({len(all_cos):,} ราย)")
        if p < max_pages:
            has_next = await _click_next(page)
            if not has_next:
                print("  ⚠️  ไม่พบปุ่ม Next — หยุดที่หน้านี้")
                break
        await asyncio.sleep(DBD_DELAY + random.uniform(0, 0.4))

    print(f"\n  ✅ Scrape เสร็จ: {len(all_cos):,} บริษัท")
    return all_cos


# ══════════════════════════════════════════
# ENRICH DISTRICT FROM PROFILE PAGE
# ══════════════════════════════════════════
async def enrich_districts(ctx, companies: list[dict]) -> list[dict]:
    total = len(companies)
    remaining = sum(1 for c in companies if not c.get("เขต/อำเภอ") or not c.get("ประเภทธุรกิจย่อย"))

    print(f"\n{'═'*60}")
    print(f"  PHASE 1.5 — ดึงเขต/อำเภอ + ประเภทธุรกิจย่อยจากหน้า profile ({remaining:,} บริษัท)")
    print(f"{'═'*60}\n")

    page = await ctx.new_page()
    for i, company in enumerate(companies, 1):
        if company.get("เขต/อำเภอ") and company.get("ประเภทธุรกิจย่อย"):
            continue

        reg = company.get("เลขทะเบียนนิติบุคคล", company.get("เลขทะเบียน", ""))
        name = company.get("ชื่อนิติบุคคล", company.get("ชื่อบริษัท", ""))
        if not reg:
            company["เขต/อำเภอ"] = ""
            company["ประเภทธุรกิจย่อย"] = classify_food_business(
                name,
                company.get("ชื่อประเภทธุรกิจ", ""),
                "",
            )
            continue

        print(f"  [{i}/{total}] {name[:50]:<50}", end="  ", flush=True)
        district = ""
        subtype = ""
        last_error = ""
        for attempt in range(3):
            try:
                await page.goto(_profile_url(reg), wait_until="domcontentloaded", timeout=30000)
                text = await _profile_body_text(page)
                address = _extract_head_office_address(text)
                district = _extract_district(address)
                objective = _extract_objectives(text)
                subtype = classify_food_business(
                    name,
                    company.get("ชื่อประเภทธุรกิจ", ""),
                    objective,
                )
                if district and subtype:
                    break
            except Exception as e:
                last_error = str(e)
            await page.wait_for_timeout(1000)

        company["เขต/อำเภอ"] = district
        company["ประเภทธุรกิจย่อย"] = subtype or classify_food_business(
            name,
            company.get("ชื่อประเภทธุรกิจ", ""),
            "",
        )
        if district:
            print(f"{district} | {company['ประเภทธุรกิจย่อย']}")
        elif last_error:
            print(f"⚠️  {last_error}")
        else:
            print(f"— | {company['ประเภทธุรกิจย่อย']}")

        if i % SAVE_EVERY == 0:
            _save_json(companies, f"checkpoint_district_{i}.json")

        await asyncio.sleep(PROFILE_DELAY + random.uniform(0, 0.3))

    await page.close()
    found = sum(1 for c in companies if c.get("เขต/อำเภอ"))
    classified = sum(1 for c in companies if c.get("ประเภทธุรกิจย่อย"))
    print(f"\n  ✅ ดึงเขต/อำเภอเสร็จ: {found:,}/{total:,}")
    print(f"  ✅ จัดประเภทธุรกิจย่อยเสร็จ: {classified:,}/{total:,}\n")
    return companies


# ══════════════════════════════════════════
# PHASE 2 — APIFY GOOGLE SEARCH  (NEW)
# ══════════════════════════════════════════
async def google_search_apify(client: httpx.AsyncClient, name: str, query_count: list) -> dict:
    """
    ค้นหาเบอร์โทรผ่าน Apify Google Search scraper
    ไม่ใช้ browser → ไม่มี CAPTCHA เด็ดขาด
    query_count เป็น list[int] เพื่อ track จำนวน queries ที่ใช้ไป (mutable reference)
    """
    result = {"เบอร์โทร": "", "เว็บไซต์": ""}
    clean = re.sub(r"\s*(บริษัท|จำกัด|จำกัด\s*\(มหาชน\)|หจก\.?)\s*", " ", name).strip()

    queries = [f'"{name}" เบอร์โทร', f'{clean} ติดต่อ โทรศัพท์']

    for query in queries:
        if query_count[0] >= MAX_GOOGLE_PER_SESSION:
            print(f"\n  ⚠️  ถึง quota {MAX_GOOGLE_PER_SESSION} queries/session — หยุด Google search")
            return result

        try:
            # Apify Google Search Scraper actor
            # https://apify.com/apify/google-search-scraper
            actor_input = {
                "queries": query,
                "maxPagesPerQuery": 1,
                "resultsPerPage": 10,
                "languageCode": "th",
                "mobileResults": False,
            }

            # Start actor run
            r = await client.post(
                "https://api.apify.com/v2/acts/apify~google-search-scraper/runs",
                params={"token": APIFY_API_TOKEN},
                json=actor_input,
                timeout=30,
            )
            query_count[0] += 1

            if r.status_code == 401:
                print(f"\n  ⚠️  Apify token ไม่ถูกต้อง (401)")
                return result

            if r.status_code == 429:
                print(f"\n  ⚠️  Apify rate limit (429) — หยุด")
                return result

            if r.status_code != 201:
                print(f" ⚠️  Apify error {r.status_code}")
                await asyncio.sleep(2)
                continue

            run_data = r.json()["data"]
            run_id = run_data["id"]
            default_dataset_id = run_data["defaultDatasetId"]

            # Wait for run to finish (poll status)
            for _ in range(30):  # timeout 60s
                await asyncio.sleep(2)
                status_r = await client.get(
                    f"https://api.apify.com/v2/acts/apify~google-search-scraper/runs/{run_id}",
                    params={"token": APIFY_API_TOKEN},
                    timeout=15,
                )
                status = status_r.json()["data"]["status"]
                if status in ("SUCCEEDED", "FAILED", "ABORTED"):
                    break

            if status != "SUCCEEDED":
                print(f" ⚠️  run {status}")
                continue

            # Get results from dataset
            dataset_r = await client.get(
                f"https://api.apify.com/v2/datasets/{default_dataset_id}/items",
                params={"token": APIFY_API_TOKEN, "format": "json"},
                timeout=15,
            )

            if dataset_r.status_code != 200:
                print(f" ⚠️  dataset error {dataset_r.status_code}")
                continue

            items = dataset_r.json()

            if not items:
                continue

            # Extract phones and websites from organic results
            full_text = ""
            for item in items:
                for org in item.get("organicResults", []):
                    title = org.get("title", "")
                    desc = org.get("description", "")
                    url = org.get("url", "")

                    # Skip blocked domains
                    if any(bd in url for bd in BLOCKED_DOMAINS):
                        continue

                    full_text += f" {title} {desc}"

                    # Find website
                    if not result["เว็บไซต์"] and any(x in url for x in [".co.th", ".com/", ".th/", ".net/"]):
                        result["เว็บไซต์"] = url[:120]

            phones = extract_phones(full_text)
            if phones:
                result["เบอร์โทร"] = phones[0]
                break

        except httpx.TimeoutException:
            print(f" ⚠️  timeout")
            await asyncio.sleep(3)
        except Exception as e:
            print(f" ⚠️  {e}")
            await asyncio.sleep(2)

        await asyncio.sleep(random.uniform(GOOGLE_MIN, GOOGLE_MAX))

    return result


async def enrich_google(companies: list[dict]) -> list[dict]:
    """Phase 2 — ใช้ Apify แทน browser ทั้งหมด"""

    # ตรวจสอบ config — ถ้าไม่มี token ให้ข้าม Phase 2
    if not APIFY_API_TOKEN:
        print("\n  ❌ ยังไม่ได้ตั้งค่า APIFY_API_TOKEN — ข้าม Phase 2 (Google lookup)")
        print("     1. สมัครที่ https://console.apify.com/sign-up")
        print("     2. คัดลอก Personal API token จาก Settings → Integrations")
        print("     3. ตั้งค่าเป็น environment variable: export APIFY_API_TOKEN=your_token")
        print("        หรือคัดลอกค่าไปใส่ในไฟล์ .env (ดู .env.example)")
        return companies

    total = len(companies)
    start = next((i for i, c in enumerate(companies) if not c.get("เบอร์โทร")), total)
    remaining = total - start

    print(f"\n{'═'*60}")
    print(f"  PHASE 2 — Apify Google Search ({remaining:,} บริษัท)")
    print(f"  quota สูงสุด session นี้: {MAX_GOOGLE_PER_SESSION} queries")
    print(f"{'═'*60}\n")

    query_count = [0]  # mutable counter shared across calls

    async with httpx.AsyncClient() as client:
        for i in range(start, total):
            name = companies[i].get("ชื่อนิติบุคคล", companies[i].get("ชื่อบริษัท", ""))
            if not name:
                continue

            print(f"  [{i+1}/{total}] {name[:55]:<55}", end="  ", flush=True)
            info = await google_search_apify(client, name, query_count)
            companies[i]["เบอร์โทร"] = info["เบอร์โทร"]
            companies[i]["เว็บไซต์"] = info["เว็บไซต์"]
            print(f"📞 {info['เบอร์โทร'] or '—'}  (query #{query_count[0]})")

            if (i + 1) % 20 == 0:
                _save_json(companies, f"checkpoint_google_{i+1}.json")

            # หยุดถ้า quota ใกล้เต็ม
            if query_count[0] >= MAX_GOOGLE_PER_SESSION:
                print(f"\n  ⚠️  หยุดที่ {query_count[0]} queries — resume ได้จาก checkpoint")
                _save_json(companies, f"checkpoint_google_{i+1}.json")
                break

            await asyncio.sleep(random.uniform(GOOGLE_MIN, GOOGLE_MAX))

    found = sum(1 for c in companies if c.get("เบอร์โทร"))
    print(f"\n  ✅ Google เสร็จ: พบเบอร์ {found:,}/{total:,} ({found/total*100:.1f}%)")
    print(f"  📊 ใช้ไป {query_count[0]} queries")
    return companies


# ══════════════════════════════════════════
# EXPORT EXCEL  (unchanged)
# ══════════════════════════════════════════
def save_excel(companies: list[dict], keyword: str) -> str:
    if not companies:
        print("⚠️  ไม่มีข้อมูล")
        return ""

    df = pd.DataFrame(companies)
    order = [
        "เลขทะเบียนนิติบุคคล",
        "ชื่อนิติบุคคล",
        "ประเภทนิติบุคคล",
        "สถานะ",
        "รหัสประเภทธุรกิจ",
        "ชื่อประเภทธุรกิจ",
        "ประเภทธุรกิจย่อย",
        "จังหวัด",
        "เขต/อำเภอ",
        "ทุนจดทะเบียน (บาท)",
        "รายได้รวม (บาท)",
        "กำไร (ขาดทุน) สุทธิ (บาท)",
        "สินทรัพย์รวม (บาท)",
        "ส่วนของผู้ถือหุ้น (บาท)",
    ]
    df = df[[c for c in order if c in df.columns]]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"dbd_{keyword}_{ts}.xlsx"

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="บริษัท")
        ws = writer.sheets["บริษัท"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        bd = Border(
            left=Side("thin", color="D0D0D0"), right=Side("thin", color="D0D0D0"),
            top=Side("thin", color="D0D0D0"), bottom=Side("thin", color="D0D0D0"),
        )
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bd
        ws.row_dimensions[1].height = 28

        even = PatternFill("solid", fgColor="EBF3FB")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = bd
                cell.alignment = Alignment(vertical="center")
                if cell.row % 2 == 0:
                    cell.fill = even

        widths = {
            "เลขทะเบียนนิติบุคคล": 20,
            "ชื่อนิติบุคคล": 45,
            "ประเภทนิติบุคคล": 22,
            "สถานะ": 22,
            "รหัสประเภทธุรกิจ": 20,
            "ชื่อประเภทธุรกิจ": 50,
            "ประเภทธุรกิจย่อย": 18,
            "จังหวัด": 18,
            "เขต/อำเภอ": 18,
            "ทุนจดทะเบียน (บาท)": 22,
            "รายได้รวม (บาท)": 22,
            "กำไร (ขาดทุน) สุทธิ (บาท)": 24,
            "สินทรัพย์รวม (บาท)": 22,
            "ส่วนของผู้ถือหุ้น (บาท)": 24,
        }
        for ci, col in enumerate(df.columns, 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = widths.get(col, 18)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    found_p = sum(1 for c in companies if c.get("เบอร์โทร"))
    found_w = sum(1 for c in companies if c.get("เว็บไซต์"))
    print(f"\n{'═'*60}")
    print(f"  ✅ บันทึก: {fname}")
    print(f"  📊 รวม:         {len(companies):,} บริษัท")
    print(f"  📞 พบเบอร์:     {found_p:,}  ({found_p/len(companies)*100:.1f}%)")
    print(f"  🌐 พบเว็บไซต์:  {found_w:,}  ({found_w/len(companies)*100:.1f}%)")
    print(f"{'═'*60}\n")
    return fname


# ══════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════
def _save_json(data, fname):
    with open(fname, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _load_json(fname):
    with open(fname, "r", encoding="utf-8") as f:
        return json.load(f)


# ══════════════════════════════════════════
# CLI
# ══════════════════════════════════════════
def ask_config() -> dict | None:
    print("\n╔══════════════════════════════════════════════════════╗")
    print("║  DBD Full Scraper v7  (datawarehouse.dbd.go.th)      ║")
    print("║  Phase 2: Apify Google Search — no CAPTCHA           ║")
    print("╚══════════════════════════════════════════════════════╝\n")

    kw = input("กรอก รหัส TSIC เช่น 56101 : ").strip()
    if not kw:
        print("ต้องกรอก keyword")
        return None

    max_p = input("ดึงกี่หน้า? (Enter = ทั้งหมด) : ").strip()
    max_pages = int(max_p) if max_p.isdigit() else None

    do_google = input("ค้น Google เบอร์โทรด้วยไหม? [y/n] default=y : ").strip().lower() != "n"

    def checkpoint_page(path: Path) -> int:
        m = re.search(r"_p(\d+)\.json$", path.name)
        return int(m.group(1)) if m else -1

    cps = sorted(Path(".").glob(f"checkpoint_{kw}_*.json"), key=checkpoint_page)
    resume = None
    if cps:
        print("\nพบ checkpoint:")
        for i, c in enumerate(cps):
            data = _load_json(str(c))
            print(f"  [{i}] {c.name}  ({len(data):,} ราย)")
        if input("Resume จาก checkpoint? [y/n] default=n : ").strip().lower() == "y":
            idx = int(input(f"  เลือก [0-{len(cps)-1}]: ").strip() or "0")
            resume = str(cps[idx])

    profile = _default_chrome_profile()
    print(f"\n⚙️  Chrome profile: {profile}")
    print(f"⚙️  keyword={kw}  หน้า={'ทั้งหมด' if not max_pages else max_pages}  Google={'✓' if do_google else '✗'}")
    if input("เริ่มเลยไหม? [y/n] : ").strip().lower() != "y":
        return None

    return dict(keyword=kw, max_pages=max_pages, do_google=do_google, resume=resume, profile=profile)


# ══════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════
async def main():
    cfg = ask_config()
    if not cfg:
        return

    async with async_playwright() as pw:
        launch_kwargs = dict(
            headless=False,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled", "--window-size=1280,900"],
        )

        # ใช้ real Chrome ถ้าหาเจอ (เพื่อ cookies / trust history)
        exe = _chrome_exe()
        if exe:
            launch_kwargs["executable_path"] = exe
            print(f"  🌐 ใช้ Chrome จริง: {exe}")
        else:
            print("  🌐 ใช้ Playwright Chromium (Chrome ไม่พบในเครื่อง)")

        # launch_persistent_context โหลด Chrome profile จริง → cookies / Google trust
        # ถ้าไม่พบ profile → fallback เป็น fresh context
        profile_path = cfg["profile"]
        if Path(profile_path).exists():
            print(f"  🌐 ใช้ Chrome profile: {profile_path}")
            ctx = await pw.chromium.launch_persistent_context(
                user_data_dir=profile_path,
                executable_path=exe,          # None = ใช้ Playwright Chromium
                headless=False,
                args=["--no-sandbox", "--disable-blink-features=AutomationControlled", "--window-size=1280,900"],
                viewport={"width": 1280, "height": 900},
                locale="th-TH",
            )
        else:
            print(f"  ⚠️  ไม่พบ Chrome profile ที่ {profile_path} — ใช้ fresh context")
            browser = await pw.chromium.launch(**launch_kwargs)
            ctx = await browser.new_context(
                viewport={"width": 1280, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale="th-TH",
            )

        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        )
        page = await ctx.new_page()

        if cfg["resume"]:
            print(f"\n📂 โหลด checkpoint: {cfg['resume']}")
            companies = _load_json(cfg["resume"])
            before_filter = len(companies)
            companies = filter_active_companies(companies)
            print(f"   โหลดได้ {before_filter:,} ราย | หลังกรองสถานะ active เหลือ {len(companies):,} ราย")
            m = re.search(r"_p(\d+)\.json$", cfg["resume"])
            resume_page = int(m.group(1)) if m else 0
            await navigate_to_list(page, cfg["keyword"])
            if resume_page > 0:
                companies = await scrape_dbd_all(
                    page,
                    cfg["keyword"],
                    max_pages=cfg["max_pages"],
                    start_page=resume_page + 1,
                    initial_companies=companies,
                )
                raw_file = f"dbd_raw_{cfg['keyword']}.json"
                _save_json(companies, raw_file)
                print(f"\n💾 บันทึกข้อมูล DBD ดิบ → {raw_file}")
        else:
            await navigate_to_list(page, cfg["keyword"])
            companies = await scrape_dbd_all(page, cfg["keyword"], max_pages=cfg["max_pages"])
            raw_file = f"dbd_raw_{cfg['keyword']}.json"
            _save_json(companies, raw_file)
            print(f"\n💾 บันทึกข้อมูล DBD ดิบ → {raw_file}")

        if companies:
            companies = await enrich_districts(ctx, companies)
            _save_json(companies, f"dbd_with_district_{cfg['keyword']}.json")

        await ctx.close()   # ปิด context (ใช้ได้ทั้ง persistent และ fresh)

    if not companies:
        print("❌ ไม่พบข้อมูลใดๆ")
        return

    if cfg["do_google"]:
        companies = await enrich_google(companies)   # pure API — ไม่ใช้ browser

    save_excel(companies, cfg["keyword"])
    _save_json(companies, f"dbd_final_{cfg['keyword']}.json")
    print(f"💾 JSON สำรอง → dbd_final_{cfg['keyword']}.json")


if __name__ == "__main__":
    asyncio.run(main())
